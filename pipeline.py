from __future__ import annotations

import ast
import argparse
import json
import os
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Set, Tuple

import pandas as pd
import requests
from mistralai.client import Mistral
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class CellCoordinate:
    sheet_name: str
    excel_row: int
    excel_col: int

    @property
    def excel_col_letter(self) -> str:
        return get_column_letter(self.excel_col)


@dataclass
class ExtractedTable:
    table_name: str
    sheet_name: str
    dataframe: pd.DataFrame
    coordinate_map: Dict[Tuple[int, str], CellCoordinate]
    coordinate_dataframe: pd.DataFrame
    header_map: Dict[str, CellCoordinate]
    bounding_box: Tuple[int, int, int, int]


class SpreadsheetParser:
    """
    Detects disconnected table-like regions in scattered worksheets and
    preserves the original Excel coordinates for every DataFrame value.
    """

    def parse_workbook(self, workbook_path: str | Path) -> Dict[str, ExtractedTable]:
        workbook = load_workbook(filename=workbook_path, data_only=True)
        extracted_tables: Dict[str, ExtractedTable] = {}

        for worksheet in workbook.worksheets:
            for table_index, component in enumerate(self._find_connected_components(worksheet)):
                table = self._component_to_table(worksheet, table_index, component)
                if table is not None and not table.dataframe.empty:
                    extracted_tables[table.table_name] = table

        if not extracted_tables:
            raise ValueError(f"No table-like regions found in workbook: {workbook_path}")

        return extracted_tables

    def _find_connected_components(self, worksheet) -> List[Set[Tuple[int, int]]]:
        populated_cells = {
            (cell.row, cell.column)
            for row in worksheet.iter_rows()
            for cell in row
            if cell.value is not None and str(cell.value).strip() != ""
        }

        visited: Set[Tuple[int, int]] = set()
        components: List[Set[Tuple[int, int]]] = []
        neighbors = (
            (-1, -1), (-1, 0), (-1, 1),
            (0, -1),           (0, 1),
            (1, -1),  (1, 0),  (1, 1),
        )

        for start in populated_cells:
            if start in visited:
                continue

            stack = [start]
            component: Set[Tuple[int, int]] = set()

            while stack:
                current = stack.pop()
                if current in visited:
                    continue
                visited.add(current)
                component.add(current)

                row, col = current
                for row_delta, col_delta in neighbors:
                    candidate = (row + row_delta, col + col_delta)
                    if candidate in populated_cells and candidate not in visited:
                        stack.append(candidate)

            if len(component) >= 4:
                components.append(component)

        return components

    def _component_to_table(
        self,
        worksheet,
        table_index: int,
        component: Set[Tuple[int, int]],
    ) -> Optional[ExtractedTable]:
        sheet_name = worksheet.title
        rows = [row for row, _ in component]
        cols = [col for _, col in component]
        min_row, max_row = min(rows), max(rows)
        min_col, max_col = min(cols), max(cols)

        grid_rows = []
        for excel_row in range(min_row, max_row + 1):
            row_values = []
            for excel_col in range(min_col, max_col + 1):
                if (excel_row, excel_col) in component:
                    row_values.append((excel_row, excel_col))
                else:
                    row_values.append(None)
            grid_rows.append(row_values)

        header_offset = self._find_header_row(grid_rows)
        if header_offset is None or header_offset >= len(grid_rows) - 1:
            return None

        header_cells = grid_rows[header_offset]
        active_columns = [
            offset
            for offset, cell_ref in enumerate(header_cells)
            if cell_ref is not None
        ]
        if not active_columns:
            return None

        headers: List[str] = []
        header_map: Dict[str, CellCoordinate] = {}
        seen_headers: Dict[str, int] = {}
        values_by_row: List[List[Any]] = []
        coordinates_by_row: List[List[Optional[CellCoordinate]]] = []
        index_labels: List[int] = []
        coordinate_map: Dict[Tuple[int, str], CellCoordinate] = {}

        value_lookup = self._build_component_value_lookup(worksheet, component)

        for column_offset in active_columns:
            cell_ref = header_cells[column_offset]
            raw_header = value_lookup[cell_ref] if cell_ref is not None else None
            header = self._normalize_header(raw_header, column_offset, seen_headers)
            headers.append(header)
            if cell_ref is not None:
                header_map[header] = CellCoordinate(sheet_name, cell_ref[0], cell_ref[1])

        for relative_row in range(header_offset + 1, len(grid_rows)):
            excel_refs = grid_rows[relative_row]
            row_payload: List[Any] = []
            row_coordinates: List[Optional[CellCoordinate]] = []
            row_has_value = False

            for header, column_offset in zip(headers, active_columns):
                cell_ref = excel_refs[column_offset]
                value = value_lookup.get(cell_ref) if cell_ref is not None else None
                row_payload.append(value)
                coordinate = (
                    CellCoordinate(sheet_name=sheet_name, excel_row=cell_ref[0], excel_col=cell_ref[1])
                    if cell_ref is not None
                    else None
                )
                row_coordinates.append(coordinate)

                if value is not None and str(value).strip() != "":
                    row_has_value = True
                    if coordinate is not None:
                        coordinate_map[(len(index_labels), header)] = coordinate

            if row_has_value:
                values_by_row.append(row_payload)
                coordinates_by_row.append(row_coordinates)
                index_labels.append(len(index_labels))

        if not values_by_row:
            return None

        dataframe = pd.DataFrame(values_by_row, columns=headers, index=index_labels)
        coordinate_dataframe = pd.DataFrame(coordinates_by_row, columns=headers, index=index_labels)
        dataframe = dataframe.dropna(axis=1, how="all")
        coordinate_dataframe = coordinate_dataframe[dataframe.columns]

        valid_columns = list(dataframe.columns)
        coordinate_map = {
            (row_index, column_name): coord
            for (row_index, column_name), coord in coordinate_map.items()
            if column_name in valid_columns
        }
        header_map = {column_name: coord for column_name, coord in header_map.items() if column_name in valid_columns}

        table_name = f"{sheet_name}_table_{table_index}"
        return ExtractedTable(
            table_name=table_name,
            sheet_name=sheet_name,
            dataframe=dataframe,
            coordinate_map=coordinate_map,
            coordinate_dataframe=coordinate_dataframe,
            header_map=header_map,
            bounding_box=(min_row, min_col, max_row, max_col),
        )

    def _build_component_value_lookup(self, worksheet, component: Set[Tuple[int, int]]) -> Dict[Tuple[int, int], Any]:
        return {
            (row, col): worksheet.cell(row=row, column=col).value
            for row, col in component
        }

    def _find_header_row(self, grid_rows: Sequence[Sequence[Optional[Tuple[int, int]]]]) -> Optional[int]:
        best_row: Optional[int] = None
        best_score = -1

        for row_index, row in enumerate(grid_rows):
            non_empty = sum(cell is not None for cell in row)
            if non_empty < 2:
                continue
            next_rows_have_data = any(
                any(candidate is not None for candidate in candidate_row)
                for candidate_row in grid_rows[row_index + 1:]
            )
            if next_rows_have_data and non_empty > best_score:
                best_row = row_index
                best_score = non_empty

        return best_row

    def _normalize_header(
        self,
        raw_value: Any,
        column_offset: int,
        seen_headers: Dict[str, int],
    ) -> str:
        base = str(raw_value).strip() if raw_value is not None and str(raw_value).strip() else f"column_{column_offset}"
        base = re.sub(r"\s+", "_", base)
        seen_headers[base] = seen_headers.get(base, 0) + 1
        if seen_headers[base] == 1:
            return base
        return f"{base}_{seen_headers[base]}"

    def parse(self, workbook_path: str | Path) -> Dict[str, ExtractedTable]:
        return self.parse_workbook(workbook_path)


class SafeCodeExecutor:
    """
    Validates a narrow subset of Python AST before executing LLM-produced
    pandas code in a restricted environment.
    """

    ALLOWED_NODES = {
        ast.Module,
        ast.Assign,
        ast.Expr,
        ast.Name,
        ast.Load,
        ast.Store,
        ast.Subscript,
        ast.Attribute,
        ast.Constant,
        ast.List,
        ast.Tuple,
        ast.Dict,
        ast.Call,
        ast.keyword,
        ast.Compare,
        ast.Eq,
        ast.NotEq,
        ast.Gt,
        ast.GtE,
        ast.Lt,
        ast.LtE,
        ast.BoolOp,
        ast.And,
        ast.Or,
        ast.BinOp,
        ast.Add,
        ast.Sub,
        ast.Mult,
        ast.Div,
        ast.Mod,
        ast.UnaryOp,
        ast.USub,
        ast.UAdd,
        ast.Not,
        ast.Slice,
        ast.Index,
    }

    ALLOWED_NAMES = {
        "tables",
        "pd",
        "RESULT",
        "int",
        "float",
        "str",
        "len",
        "True",
        "False",
        "None",
    }

    BLOCKED_ATTRIBUTE_PREFIXES = ("__",)
    BLOCKED_NAMES = {"eval", "exec", "open", "compile", "globals", "locals", "__import__"}
    ALLOWED_ATTRIBUTES = {
        "loc",
        "iloc",
        "at",
        "iat",
        "index",
        "columns",
        "values",
        "shape",
        "dtype",
        "dtypes",
        "tolist",
        "sum",
        "mean",
        "min",
        "max",
        "idxmin",
        "idxmax",
        "sort_values",
        "sort_index",
        "groupby",
        "agg",
        "aggregate",
        "reset_index",
        "set_index",
        "merge",
        "join",
        "query",
        "isin",
        "isna",
        "notna",
        "fillna",
        "dropna",
        "astype",
        "head",
        "tail",
        "str",
        "contains",
        "startswith",
        "endswith",
        "lower",
        "upper",
        "item",
    }

    def validate(self, code: str) -> None:
        tree = ast.parse(code, mode="exec")
        assigned_names = {
            node.id
            for node in ast.walk(tree)
            if isinstance(node, ast.Name) and isinstance(node.ctx, ast.Store)
        }

        for node in ast.walk(tree):
            if type(node) not in self.ALLOWED_NODES:
                raise ValueError(f"Blocked AST node: {type(node).__name__}")

            if isinstance(node, ast.Name):
                if node.id in self.BLOCKED_NAMES:
                    raise ValueError(f"Blocked name: {node.id}")
                if node.id not in self.ALLOWED_NAMES and node.id not in assigned_names:
                    raise ValueError(
                        f"Unexpected symbol '{node.id}'. "
                        "Restrict generated code to tables, pd, RESULT, and basic casts."
                    )

            if isinstance(node, ast.Attribute):
                if any(node.attr.startswith(prefix) for prefix in self.BLOCKED_ATTRIBUTE_PREFIXES):
                    raise ValueError(f"Blocked attribute access: {node.attr}")
                if node.attr not in self.ALLOWED_ATTRIBUTES:
                    raise ValueError(f"Attribute '{node.attr}' is not allowed in the safe execution sandbox.")

    def execute(self, code: str, tables: Dict[str, pd.DataFrame]) -> Dict[str, Any]:
        self.validate(code)

        sandbox_globals = {
            "__builtins__": {},
            "pd": pd,
            "tables": tables,
            "int": int,
            "float": float,
            "str": str,
            "len": len,
            "True": True,
            "False": False,
            "None": None,
        }
        sandbox_locals: Dict[str, Any] = {}
        exec(code, sandbox_globals, sandbox_locals)

        if "RESULT" not in sandbox_locals:
            raise ValueError("Generated code must assign the final payload to RESULT.")

        result = sandbox_locals["RESULT"]
        required_keys = {"value", "table_name", "row_index", "column_name"}
        if not isinstance(result, dict) or not required_keys.issubset(result):
            raise ValueError(
                "RESULT must be a dict with keys: value, table_name, row_index, column_name."
            )

        return result


class TableLLMInference:
    """
    Uses a local TableLLM-style endpoint when available. For offline demos,
    a deterministic fallback code generator is included.
    """

    def __init__(
        self,
        endpoint_url: Optional[str] = None,
        model_name: str = "tablellm-local",
        timeout_seconds: int = 60,
    ) -> None:
        self.endpoint_url = endpoint_url
        self.model_name = model_name
        self.timeout_seconds = timeout_seconds
        self.executor = SafeCodeExecutor()

    def answer_query(
        self,
        tables: Dict[str, ExtractedTable],
        query: str,
    ) -> Dict[str, Any]:
        prompt = self._build_prompt(tables, query)
        generated_code = self._infer_code(prompt, tables, query)
        dataframe_map = {name: table.dataframe for name, table in tables.items()}
        result = self.executor.execute(generated_code, dataframe_map)
        result["generated_code"] = generated_code
        return result

    def _build_prompt(self, tables: Dict[str, ExtractedTable], query: str) -> str:
        table_context = []
        for table_name, table in tables.items():
            preview = table.dataframe.head(5).to_dict(orient="records")
            table_context.append(
                {
                    "table_name": table_name,
                    "columns": list(table.dataframe.columns),
                    "shape": list(table.dataframe.shape),
                    "preview": preview,
                }
            )

        instructions = {
            "task": "Write pandas-only Python code that answers the query.",
            "result_contract": {
                "variable_name": "RESULT",
                "format": {
                    "value": "answer value",
                    "table_name": "source table name",
                    "row_index": "integer pandas row index",
                    "column_name": "source column name",
                },
            },
            "constraints": [
                "Use the provided `tables` dictionary only.",
                "Do not import modules.",
                "Reference explicit row indices and a concrete source column in RESULT.",
                "Return only Python code.",
            ],
            "query": query,
            "tables": table_context,
        }
        return json.dumps(instructions, indent=2)

    def _infer_code(
        self,
        prompt: str,
        tables: Dict[str, ExtractedTable],
        query: str,
    ) -> str:
        if self.endpoint_url:
            try:
                response = requests.post(
                    self.endpoint_url,
                    json={
                        "model": self.model_name,
                        "prompt": prompt,
                        "temperature": 0,
                    },
                    timeout=self.timeout_seconds,
                )
                response.raise_for_status()
                payload = response.json()
                code = payload.get("code") or payload.get("response") or payload.get("text")
                if not code:
                    raise ValueError("Local TableLLM endpoint returned no code field.")
                return self._strip_code_fences(code.strip())
            except Exception as exc:
                print(f"[WARN] Falling back to deterministic demo inference: {exc}")

        return self._deterministic_demo_code(tables, query)

    def _strip_code_fences(self, code: str) -> str:
        if code.startswith("```"):
            code = re.sub(r"^```[a-zA-Z0-9_]*\n", "", code)
            code = re.sub(r"\n```$", "", code)
        return code.strip()

    def _deterministic_demo_code(
        self,
        tables: Dict[str, ExtractedTable],
        query: str,
    ) -> str:
        lowered_query = query.lower()

        for table_name, table in tables.items():
            columns = {str(column).lower(): str(column) for column in table.dataframe.columns}
            if {"region", "q2_sales"}.issubset(columns):
                region_match = re.search(r"\b(north|south|east|west)\b", lowered_query)
                if region_match and ("q2" in lowered_query or "quarter 2" in lowered_query):
                    region_value = region_match.group(1).title()
                    region_col = columns["region"]
                    q2_col = columns["q2_sales"]
                    return f"""
row_idx = tables["{table_name}"].index[tables["{table_name}"]["{region_col}"] == "{region_value}"][0]
value = tables["{table_name}"].loc[row_idx, "{q2_col}"]
RESULT = {{"value": value, "table_name": "{table_name}", "row_index": int(row_idx), "column_name": "{q2_col}"}}
""".strip()

        raise ValueError(
            "Fallback inference could not map the query. "
            "Provide a local TableLLM endpoint or adjust the demo query."
        )


class ProvenanceMapper:
    """
    Converts DataFrame row/column references back to exact Excel coordinates.
    """

    def map_result(self, result: Dict[str, Any], tables: Dict[str, ExtractedTable]) -> str:
        provenance = self.get_result_provenance(result, tables)
        return (
            f"{result['value']}\n"
            f"Source: Sheet[{provenance['sheet_name']}], Row [{provenance['excel_row']}], Col [{provenance['excel_col']}]"
        )

    def get_result_provenance(self, result: Dict[str, Any], tables: Dict[str, ExtractedTable]) -> Dict[str, Any]:
        table_name = result["table_name"]
        row_index = int(result["row_index"])
        column_name = str(result["column_name"])

        if table_name not in tables:
            raise KeyError(f"Unknown table_name in RESULT: {table_name}")

        table = tables[table_name]
        coordinate = table.coordinate_map.get((row_index, column_name))
        if coordinate is None:
            raise KeyError(
                f"No Excel coordinate found for table={table_name}, row_index={row_index}, column={column_name}"
            )

        return {
            "table_name": table_name,
            "pandas_row_index": row_index,
            "column_name": column_name,
            "sheet_name": coordinate.sheet_name,
            "excel_row": coordinate.excel_row,
            "excel_col": coordinate.excel_col,
            "excel_col_letter": coordinate.excel_col_letter,
            "cell_address": f"{coordinate.excel_col_letter}{coordinate.excel_row}",
        }


class GemmaQueryManager:
    """
    Uses a local Ollama-served Gemma 4 model to convert free-form user queries
    into a structured lookup intent that downstream resolvers can execute.
    """

    def __init__(
        self,
        endpoint_url: Optional[str] = "http://127.0.0.1:11434/api/chat",
        model_name: str = "gemma4:e2b",
        timeout_seconds: int = 60,
        enabled: bool = True,
    ) -> None:
        self.endpoint_url = endpoint_url
        self.model_name = model_name
        self.timeout_seconds = timeout_seconds
        self.enabled = enabled

    def parse_query(self, query: str) -> Optional[Dict[str, Any]]:
        if not self.enabled or not self.endpoint_url:
            return None

        system_prompt = (
            "You are a query manager for spreadsheet retrieval. "
            "Convert the user's question into compact JSON only. "
            "Return an object with keys: "
            "`query_type`, `metric`, `category`, `filters`, `aggregation`, `notes`. "
            "Use query_type=`lookup` for direct metric lookups. "
            "If a field is unknown, use null. "
            "Do not include markdown, prose, or code fences."
        )

        response = requests.post(
            self.endpoint_url,
            json={
                "model": self.model_name,
                "stream": False,
                "messages": [
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": query},
                ],
                "options": {"temperature": 0},
            },
            timeout=self.timeout_seconds,
        )
        response.raise_for_status()
        payload = response.json()
        content = payload.get("message", {}).get("content", "")
        content = self._strip_reasoning_artifacts(content)
        parsed = self._extract_json_object(content)
        if not isinstance(parsed, dict):
            raise ValueError("Gemma query manager returned non-object JSON.")
        return parsed

    def _strip_reasoning_artifacts(self, content: str) -> str:
        content = re.sub(r"<\|channel\|>thought.*?<\|channel\|>", "", content, flags=re.DOTALL)
        content = re.sub(r"^```json\s*", "", content.strip())
        content = re.sub(r"\s*```$", "", content.strip())
        return content.strip()

    def _extract_json_object(self, content: str) -> Dict[str, Any]:
        try:
            return json.loads(content)
        except json.JSONDecodeError:
            match = re.search(r"\{.*\}", content, flags=re.DOTALL)
            if not match:
                raise
            return json.loads(match.group(0))


class MistralQueryManager:
    """
    Uses Mistral chat completions in JSON mode to normalize user queries into
    a structured lookup intent for downstream provenance-aware resolution.
    """

    def __init__(
        self,
        api_key: Optional[str] = None,
        api_key_env_var: str = "MISTRAL_API_KEY",
        model_name: str = "mistral-small-latest",
        enabled: bool = True,
    ) -> None:
        self.api_key = api_key or os.getenv(api_key_env_var)
        self.api_key_env_var = api_key_env_var
        self.model_name = model_name
        self.enabled = enabled and bool(self.api_key)
        self.client = Mistral(api_key=self.api_key) if self.enabled else None

    def parse_query(self, query: str) -> Optional[Dict[str, Any]]:
        if not self.enabled or self.client is None:
            return None

        system_prompt = (
            "You are a query manager for spreadsheet retrieval. "
            "Convert the user's question into compact JSON. "
            "Return exactly one JSON object with keys: "
            "`query_type`, `metric`, `category`, `filters`, `aggregation`, `notes`. "
            "Use query_type=`lookup` for direct metric lookups. "
            "Use null for unknown fields."
        )

        response = self.client.chat.complete(
            model=self.model_name,
            response_format={"type": "json_object"},
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": query},
            ],
        )
        content = response.choices[0].message.content
        if isinstance(content, list):
            content = "".join(
                item.get("text", "")
                for item in content
                if isinstance(item, dict)
            )
        if not isinstance(content, str):
            raise ValueError("Mistral query manager returned unsupported content format.")
        return json.loads(content)


class WorkbookHeuristicQueryResolver:
    """
    Resolves simple metric/category lookup questions directly from the workbook
    when TableLLM or table-level heuristics are unavailable.
    """

    def resolve(
        self,
        workbook_path: str | Path,
        query: str,
        structured_query: Optional[Dict[str, Any]] = None,
    ) -> Optional[Dict[str, Any]]:
        metric_phrase, category_phrase = self._extract_lookup_phrases(query, structured_query)
        workbook = load_workbook(filename=workbook_path, data_only=True)
        best_match = self._resolve_schema_lookup(workbook, query, metric_phrase, category_phrase)
        if best_match is None:
            best_match = self._resolve_phrase_lookup(workbook, metric_phrase, category_phrase) if metric_phrase else None
        if best_match is None:
            best_match = self._resolve_flexible_lookup(workbook, query)
        if best_match is None:
            return None
        best_match.pop("_score", None)
        return best_match

    def _resolve_schema_lookup(
        self,
        workbook,
        query: str,
        metric_phrase: Optional[str],
        category_phrase: Optional[str],
    ) -> Optional[Dict[str, Any]]:
        query_tokens = self._meaningful_tokens(query)
        best_match: Optional[Dict[str, Any]] = None

        for worksheet in workbook.worksheets:
            sections = self._extract_section_blocks(worksheet)
            for section in sections:
                header_tokens = self._meaningful_tokens(section["header_label"])
                header_overlap = len(query_tokens & header_tokens)

                if not section["rows"]:
                    continue

                for item_row in section["rows"]:
                    row_tokens = self._meaningful_tokens(item_row["label"])
                    row_overlap = len(query_tokens & row_tokens)

                    # Child-row lookup: category + section => use the main value column F.
                    if row_overlap > 0 and (header_overlap > 0 or category_phrase):
                        score = (header_overlap * 40) + (row_overlap * 60)
                        candidate = self._build_candidate(
                            worksheet=worksheet,
                            value_cell=item_row["value_cell"],
                            metric_cell=section["header_cell"],
                            category_cell_coordinate=item_row["label_cell"].coordinate,
                            score=score,
                        )
                        if best_match is None or candidate["_score"] > best_match["_score"]:
                            best_match = candidate

                # Section-only lookup: match the section header and prefer the H-column factor/reference.
                if header_overlap > 0 and not self._query_mentions_any_row_label(query_tokens, section["rows"]):
                    header_value_cell = section.get("header_value_cell")
                    if header_value_cell is not None:
                        score = header_overlap * 100
                        candidate = self._build_candidate(
                            worksheet=worksheet,
                            value_cell=header_value_cell,
                            metric_cell=section["header_cell"],
                            category_cell_coordinate=None,
                            score=score,
                        )
                        if best_match is None or candidate["_score"] > best_match["_score"]:
                            best_match = candidate

        return best_match

    def _extract_section_blocks(self, worksheet) -> List[Dict[str, Any]]:
        sections: List[Dict[str, Any]] = []
        current_section: Optional[Dict[str, Any]] = None

        for row_number in range(1, worksheet.max_row + 1):
            header_cell = worksheet.cell(row=row_number, column=2)
            row_label_cell = worksheet.cell(row=row_number, column=3)
            main_value_cell = worksheet.cell(row=row_number, column=6)
            unit_cell = worksheet.cell(row=row_number, column=7)
            header_factor_cell = worksheet.cell(row=row_number, column=8)

            header_label = self._string_or_none(header_cell.value)
            row_label = self._string_or_none(row_label_cell.value)

            if self._is_scope_separator(worksheet, row_number):
                current_section = None
                continue

            if self._is_section_header_row(header_label):
                current_section = {
                    "header_label": header_label,
                    "header_cell": header_cell,
                    "header_value_cell": header_factor_cell if isinstance(header_factor_cell.value, (int, float)) else None,
                    "rows": [],
                }
                sections.append(current_section)
                continue

            if current_section is not None and self._is_section_item_row(row_label, main_value_cell.value, unit_cell.value):
                current_section["rows"].append(
                    {
                        "label": row_label,
                        "label_cell": row_label_cell,
                        "value_cell": main_value_cell,
                        "unit_cell": unit_cell,
                    }
                )

        return sections

    def _is_scope_separator(self, worksheet, row_number: int) -> bool:
        first_cell = self._string_or_none(worksheet.cell(row=row_number, column=1).value)
        return bool(first_cell and first_cell.lower().startswith("scope"))

    def _is_section_header_row(self, header_label: Optional[str]) -> bool:
        if not header_label:
            return False
        lowered = header_label.strip().lower()
        if lowered.startswith("scope"):
            return False
        return len(self._meaningful_tokens(lowered)) >= 1

    def _is_section_item_row(self, row_label: Optional[str], main_value: Any, unit_value: Any) -> bool:
        if not row_label:
            return False
        return isinstance(main_value, (int, float)) and unit_value is not None

    def _query_mentions_any_row_label(self, query_tokens: Set[str], rows: List[Dict[str, Any]]) -> bool:
        for row in rows:
            if query_tokens & self._meaningful_tokens(row["label"]):
                return True
        return False

    def _string_or_none(self, value: Any) -> Optional[str]:
        if value is None:
            return None
        text = str(value).strip()
        return text if text else None

    def _resolve_phrase_lookup(
        self,
        workbook,
        metric_phrase: str,
        category_phrase: Optional[str],
    ) -> Optional[Dict[str, Any]]:
        best_match: Optional[Dict[str, Any]] = None

        for worksheet in workbook.worksheets:
            metric_hits = self._find_phrase_hits(worksheet, metric_phrase)
            category_hits = self._find_phrase_hits(worksheet, category_phrase) if category_phrase else []

            for metric_hit in metric_hits:
                value_cell = self._find_value_cell_to_right(worksheet, metric_hit)
                if value_cell is None:
                    continue

                score = self._score_match(metric_hit, category_hits)
                candidate = self._build_candidate(
                    worksheet=worksheet,
                    value_cell=value_cell,
                    metric_cell=metric_hit,
                    category_cell_coordinate=score["category_cell"],
                    score=score["score"],
                )
                if best_match is None or candidate["_score"] > best_match["_score"]:
                    best_match = candidate

        return best_match

    def _resolve_flexible_lookup(self, workbook, query: str) -> Optional[Dict[str, Any]]:
        query_tokens = self._meaningful_tokens(query)
        if not query_tokens:
            return None

        best_match: Optional[Dict[str, Any]] = None

        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    label_text = str(cell.value)
                    label_tokens = self._meaningful_tokens(label_text)
                    if not label_tokens:
                        continue

                    token_overlap = len(query_tokens & label_tokens)
                    if token_overlap == 0:
                        continue

                    candidate = self._candidate_from_contextual_value(worksheet, cell, query_tokens, token_overlap)
                    if candidate is None:
                        continue

                    if best_match is None or candidate["_score"] > best_match["_score"]:
                        best_match = candidate

        return best_match

    def _candidate_from_contextual_value(
        self,
        worksheet,
        label_cell,
        query_tokens: Set[str],
        token_overlap: int,
    ) -> Optional[Dict[str, Any]]:
        for row_offset in range(0, 4):
            candidate_row = label_cell.row + row_offset
            if candidate_row > worksheet.max_row:
                break

            row_context_tokens = self._meaningful_tokens(
                " ".join(
                    str(worksheet.cell(candidate_row, col).value)
                    for col in range(1, min(worksheet.max_column, 12) + 1)
                    if worksheet.cell(candidate_row, col).value is not None
                )
            )
            contextual_overlap = len(query_tokens & row_context_tokens)

            for value_col in range(label_cell.column + 1, min(worksheet.max_column, label_cell.column + 8) + 1):
                value_cell = worksheet.cell(candidate_row, value_col)
                if not isinstance(value_cell.value, (int, float)):
                    continue

                category_cell_coordinate = self._best_category_cell_in_row(worksheet, candidate_row, query_tokens)
                score = (token_overlap * 30) + (contextual_overlap * 20) - (row_offset * 5)
                return self._build_candidate(
                    worksheet=worksheet,
                    value_cell=value_cell,
                    metric_cell=label_cell,
                    category_cell_coordinate=category_cell_coordinate,
                    score=score,
                )

        return None

    def _best_category_cell_in_row(self, worksheet, row_number: int, query_tokens: Set[str]) -> Optional[str]:
        best_coordinate = None
        best_overlap = 0
        for col in range(1, min(worksheet.max_column, 12) + 1):
            cell = worksheet.cell(row_number, col)
            if cell.value is None or isinstance(cell.value, (int, float)):
                continue
            overlap = len(query_tokens & self._meaningful_tokens(str(cell.value)))
            if overlap > best_overlap:
                best_overlap = overlap
                best_coordinate = cell.coordinate
        return best_coordinate

    def _build_candidate(
        self,
        worksheet,
        value_cell,
        metric_cell,
        category_cell_coordinate: Optional[str],
        score: int,
    ) -> Dict[str, Any]:
        return {
            "answer": value_cell.value,
            "formatted_answer": (
                f"{value_cell.value}\n"
                f"Source: Sheet[{worksheet.title}], Row [{value_cell.row}], Col [{value_cell.column}]"
            ),
            "generated_code": None,
            "provenance": {
                "table_name": None,
                "pandas_row_index": None,
                "column_name": None,
                "sheet_name": worksheet.title,
                "excel_row": value_cell.row,
                "excel_col": value_cell.column,
                "excel_col_letter": get_column_letter(value_cell.column),
                "cell_address": value_cell.coordinate,
                "metric_cell": metric_cell.coordinate,
                "category_cell": category_cell_coordinate,
            },
            "_score": score,
        }

    def _extract_lookup_phrases(
        self,
        query: str,
        structured_query: Optional[Dict[str, Any]],
    ) -> Tuple[str, Optional[str]]:
        if structured_query:
            metric = structured_query.get("metric")
            category = structured_query.get("category")
            if metric:
                return str(metric).strip(), str(category).strip() if category else None
        return self._split_query(query)

    def _split_query(self, query: str) -> Tuple[str, Optional[str]]:
        normalized = re.sub(r"\s+", " ", query.strip().lower())
        normalized = re.sub(r"^(what is|what's|find|show me|give me)\s+", "", normalized)
        if " in " in normalized:
            metric, category = normalized.split(" in ", 1)
            return metric.strip(), category.strip()
        return normalized.strip(), None

    def _find_phrase_hits(self, worksheet, phrase: str) -> List[Any]:
        if not phrase:
            return []
        normalized_phrase = self._normalize_text(phrase)
        hits = []
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                if normalized_phrase in self._normalize_text(str(cell.value)):
                    hits.append(cell)
        return hits

    def _find_value_cell_to_right(self, worksheet, label_cell) -> Optional[Any]:
        for column in range(label_cell.column + 1, min(label_cell.column + 8, worksheet.max_column) + 1):
            candidate = worksheet.cell(row=label_cell.row, column=column)
            if isinstance(candidate.value, (int, float)):
                return candidate
        return None

    def _score_match(self, metric_cell, category_hits: List[Any]) -> Dict[str, Any]:
        if not category_hits:
            return {"score": 1, "category_cell": None}

        best_score = -10**9
        best_category_cell = None
        for category_cell in category_hits:
            row_gap = abs(metric_cell.row - category_cell.row)
            col_gap = abs(metric_cell.column - category_cell.column)
            score = -((row_gap * 10) + col_gap)
            if category_cell.row <= metric_cell.row:
                score += 15
            if category_cell.column <= metric_cell.column:
                score += 5
            if score > best_score:
                best_score = score
                best_category_cell = category_cell.coordinate
        return {"score": best_score, "category_cell": best_category_cell}

    def _normalize_text(self, text: str) -> str:
        return re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()

    def _meaningful_tokens(self, text: str) -> Set[str]:
        stopwords = {"what", "is", "the", "a", "an", "of", "for", "and", "to", "in", "use"}
        return {
            token
            for token in self._normalize_text(text).split()
            if token and token not in stopwords and len(token) > 1
        }


class QueryPipeline:
    """
    Orchestrates parsing, querying, provenance lookup, and JSON persistence.
    """

    def __init__(
        self,
        endpoint_url: Optional[str] = None,
        model_name: str = "tablellm-local",
        timeout_seconds: int = 60,
        query_manager_provider: str = "mistral",
        query_manager_endpoint: Optional[str] = "http://127.0.0.1:11434/api/chat",
        query_manager_model: str = "gemma4:e2b",
        mistral_model: str = "mistral-small-latest",
        mistral_api_key: Optional[str] = None,
        mistral_api_key_env_var: str = "MISTRAL_API_KEY",
        use_gemma_query_manager: bool = True,
    ) -> None:
        self.parser = SpreadsheetParser()
        self.tablellm = TableLLMInference(
            endpoint_url=endpoint_url,
            model_name=model_name,
            timeout_seconds=timeout_seconds,
        )
        self.provenance_mapper = ProvenanceMapper()
        self.workbook_resolver = WorkbookHeuristicQueryResolver()
        if query_manager_provider == "mistral":
            self.query_manager = MistralQueryManager(
                api_key=mistral_api_key,
                api_key_env_var=mistral_api_key_env_var,
                model_name=mistral_model,
                enabled=use_gemma_query_manager,
            )
        elif query_manager_provider == "ollama-gemma4":
            self.query_manager = GemmaQueryManager(
                endpoint_url=query_manager_endpoint,
                model_name=query_manager_model,
                timeout_seconds=timeout_seconds,
                enabled=use_gemma_query_manager,
            )
        else:
            self.query_manager = None

    def parse_workbook(self, workbook_path: str | Path) -> Dict[str, ExtractedTable]:
        return self.parser.parse(workbook_path)

    def answer_query(
        self,
        workbook_path: str | Path,
        query: str,
        output_json_path: Optional[str | Path] = None,
        history_json_path: Optional[str | Path] = None,
    ) -> Dict[str, Any]:
        tables = self.parse_workbook(workbook_path)
        structured_query: Optional[Dict[str, Any]] = None
        if self.query_manager and self.query_manager.enabled:
            try:
                structured_query = self.query_manager.parse_query(query)
            except Exception as exc:
                print(f"[WARN] Query manager unavailable, falling back to heuristic parsing: {exc}")

        try:
            raw_result = self.tablellm.answer_query(tables, query)
            provenance = self.provenance_mapper.get_result_provenance(raw_result, tables)
            formatted_answer = self.provenance_mapper.map_result(raw_result, tables)
            payload = {
                "workbook_path": str(Path(workbook_path).resolve()),
                "query": query,
                "answer": raw_result["value"],
                "formatted_answer": formatted_answer,
                "provenance": provenance,
                "resolver": "tablellm_or_table_fallback",
            }
        except ValueError:
            heuristic_result = self.workbook_resolver.resolve(workbook_path, query, structured_query=structured_query)
            if heuristic_result is None:
                raise
            payload = {
                "workbook_path": str(Path(workbook_path).resolve()),
                "query": query,
                "answer": heuristic_result["answer"],
                "formatted_answer": heuristic_result["formatted_answer"],
                "provenance": heuristic_result["provenance"],
                "resolver": "workbook_heuristic_fallback",
            }
            if structured_query is not None:
                payload["resolver"] = "llm_query_manager_plus_workbook_heuristic"

        if output_json_path is not None:
            self.save_json(payload, output_json_path)
            payload["output_json_path"] = str(Path(output_json_path).resolve())

        if history_json_path is not None:
            self.append_history(payload, history_json_path)
            payload["history_json_path"] = str(Path(history_json_path).resolve())

        return payload

    def save_json(self, payload: Dict[str, Any], output_json_path: str | Path) -> None:
        output_path = Path(output_json_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(json.dumps(payload, indent=2, default=self._json_default), encoding="utf-8")

    def append_history(self, payload: Dict[str, Any], history_json_path: str | Path) -> None:
        history_path = Path(history_json_path)
        history_path.parent.mkdir(parents=True, exist_ok=True)

        history: List[Dict[str, Any]]
        if history_path.exists():
            try:
                existing = json.loads(history_path.read_text(encoding="utf-8"))
                history = existing if isinstance(existing, list) else []
            except json.JSONDecodeError:
                history = []
        else:
            history = []

        history_entry = {
            "timestamp_utc": datetime.now(timezone.utc).isoformat(),
            "workbook_path": payload["workbook_path"],
            "query": payload["query"],
            "answer": payload["answer"],
            "formatted_answer": payload["formatted_answer"],
            "provenance": payload["provenance"],
            "resolver": payload["resolver"],
        }
        history.append(history_entry)
        history_path.write_text(json.dumps(history, indent=2, default=self._json_default), encoding="utf-8")

    def _json_default(self, value: Any) -> Any:
        if value is None:
            return None
        if isinstance(value, CellCoordinate):
            return {
                "sheet_name": value.sheet_name,
                "excel_row": value.excel_row,
                "excel_col": value.excel_col,
                "excel_col_letter": value.excel_col_letter,
                "cell_address": f"{value.excel_col_letter}{value.excel_row}",
            }
        if isinstance(value, Path):
            return str(value)
        if hasattr(value, "item"):
            try:
                return value.item()
            except Exception:
                return str(value)
        raise TypeError(f"Object of type {type(value).__name__} is not JSON serializable")


def create_dummy_scattered_workbook(output_path: str | Path) -> Path:
    output_path = Path(output_path)
    workbook = Workbook()

    sales_sheet = workbook.active
    sales_sheet.title = "SalesData"

    sales_sheet["B2"] = "Region"
    sales_sheet["C2"] = "Q1_Sales"
    sales_sheet["D2"] = "Q2_Sales"
    sales_sheet["B3"] = "North"
    sales_sheet["C3"] = 120
    sales_sheet["D3"] = 140
    sales_sheet["B4"] = "South"
    sales_sheet["C4"] = 90
    sales_sheet["D4"] = 160
    sales_sheet["B5"] = "West"
    sales_sheet["C5"] = 115
    sales_sheet["D5"] = 130

    sales_sheet["G8"] = "Product"
    sales_sheet["H8"] = "Price"
    sales_sheet["G9"] = "Widget"
    sales_sheet["H9"] = 25
    sales_sheet["G10"] = "Gadget"
    sales_sheet["H10"] = 40

    finance_sheet = workbook.create_sheet("Finance")
    finance_sheet["C3"] = "Metric"
    finance_sheet["D3"] = "Value"
    finance_sheet["C4"] = "Revenue"
    finance_sheet["D4"] = 900000
    finance_sheet["C5"] = "Cost"
    finance_sheet["D5"] = 550000

    workbook.save(output_path)
    return output_path


def build_argument_parser() -> argparse.ArgumentParser:
    argument_parser = argparse.ArgumentParser(description="Query scattered Excel sheets with provenance tracking.")
    argument_parser.add_argument("--workbook", type=Path, default=Path("dummy_scattered.xlsx"))
    argument_parser.add_argument("--query", type=str, default="What is the Q2 sales value for South?")
    argument_parser.add_argument("--output-json", type=Path, default=Path("query_result.json"))
    argument_parser.add_argument("--history-json", type=Path, default=Path("query_history.json"))
    argument_parser.add_argument("--tablellm-endpoint", type=str, default=None)
    argument_parser.add_argument("--model-name", type=str, default="tablellm-local")
    argument_parser.add_argument(
        "--query-manager-provider",
        choices=["mistral", "ollama-gemma4", "none"],
        default="mistral",
    )
    argument_parser.add_argument("--query-manager-endpoint", type=str, default="http://127.0.0.1:11434/api/chat")
    argument_parser.add_argument("--query-manager-model", type=str, default="gemma4:e2b")
    argument_parser.add_argument("--mistral-model", type=str, default="mistral-small-latest")
    argument_parser.add_argument("--mistral-api-key-env-var", type=str, default="MISTRAL_API_KEY")
    argument_parser.add_argument(
        "--disable-query-manager",
        action="store_true",
        help="Disable the LLM query manager and use only rule-based parsing.",
    )
    argument_parser.add_argument(
        "--create-demo-workbook",
        action="store_true",
        help="Create the dummy scattered workbook before querying.",
    )
    return argument_parser


def main() -> None:
    args = build_argument_parser().parse_args()

    if args.create_demo_workbook or not args.workbook.exists():
        create_dummy_scattered_workbook(args.workbook)

    pipeline = QueryPipeline(
        endpoint_url=args.tablellm_endpoint,
        model_name=args.model_name,
        query_manager_provider=args.query_manager_provider,
        query_manager_endpoint=args.query_manager_endpoint,
        query_manager_model=args.query_manager_model,
        mistral_model=args.mistral_model,
        mistral_api_key_env_var=args.mistral_api_key_env_var,
        use_gemma_query_manager=not args.disable_query_manager,
    )
    payload = pipeline.answer_query(
        workbook_path=args.workbook,
        query=args.query,
        output_json_path=args.output_json,
        history_json_path=args.history_json,
    )

    print("Final answer:")
    print(payload["formatted_answer"])
    print(f"\nStored JSON: {payload['output_json_path']}")
    print(f"Query history: {payload['history_json_path']}")


if __name__ == "__main__":
    main()
