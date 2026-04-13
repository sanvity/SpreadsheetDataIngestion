import streamlit as st
import pandas as pd
import json
import os
from openpyxl import load_workbook
import google.generativeai as genai

# Import from existing modules
from extractor import extract_table_with_coords
from execute_and_map import execute_and_augment

# --- Configuration ---
st.set_page_config(page_title="Excel Table Extractor", layout="wide")
st.title("📊 Excel Data Extraction Pipeline")

HISTORY_FILE = "query_history.json"

# --- Helper Functions ---
def load_history():
    if os.path.exists(HISTORY_FILE):
        with open(HISTORY_FILE, "r") as f:
            try:
                return json.load(f)
            except:
                return []
    return []

def save_to_history(query, value, citation):
    history = load_history()
    history.append({
        "query": query,
        "value": value if isinstance(value, dict) else str(value), # Preserve JSON structure if dict
        "citation": citation
    })
    with open(HISTORY_FILE, "w") as f:
        json.dump(history, f, indent=4)

def generate_llm_code(df, query, api_key):
    try:
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel("gemini-2.5-flash")
        
        prompt = f"""[INST]Below are the first few lines of a CSV file loaded as a pandas DataFrame named `df`.
You are tasked with writing a Python program to solve the provided question. Your code must process `df` and print EXACTLY the first matching integer DataFrame index of the row that contains the answer. 
Example to print the first matching index: `print(df[...].index[0])`
Do not print strings, do not print Pandas Index objects. ONLY PRINT THE MILLISECOND INTEGER VALUE.
Only provide the python code, nothing else. Do not use markdown backticks. 

Header and first few lines of CSV file:
{df.head(5).to_csv(index=False)}

Question: {query}[/INST]"""
        
        response = model.generate_content(prompt)
        code = response.text.replace("```python", "").replace("```", "").strip()
        return code
    except Exception as e:
        st.error(f"LLM Error: {e}")
        return None

# --- Sidebar ---
with st.sidebar:
    st.header("⚙️ Configuration")
    google_api_key = st.text_input("Gemini API Key", type="password", help="Required to dynamically generate Python logic on Streamlit Cloud.")
    
    st.header("💾 Query History (JSON)")
    history_data = load_history()
    if history_data:
        # Provide Download Button for the JSON
        st.download_button(
            label="Download JSON History",
            data=json.dumps(history_data, indent=4),
            file_name="query_history.json",
            mime="application/json",
            use_container_width=True
        )
        # Show quick preview
        with st.expander("View History", expanded=False):
            st.json(history_data)
    else:
        st.info("No queries saved yet.")

# --- Main App ---
uploaded_file = st.file_uploader("Upload an Excel Spreadsheet", type=["xlsx", "xls", "xlsm"])

if uploaded_file is not None:
    # Read available sheets
    try:
        wb = load_workbook(uploaded_file, read_only=True)
        sheet_names = wb.sheetnames
    except Exception as e:
        st.error(f"Failed to read file: {e}")
        st.stop()

    sheet_name = st.selectbox("Select Sheet", sheet_names)

    # Need to save upload to disk for openpyxl full loading mechanisms
    tmp_path = f"/tmp/streamlit_{uploaded_file.name}"
    with open(tmp_path, "wb") as f:
        f.write(uploaded_file.getbuffer())

    # Extract Data
    with st.spinner("Extracting data and mapping coordinates..."):
        try:
            wb_full = load_workbook(tmp_path)
            sheet = wb_full[sheet_name]
            
            # Using full bounds by default to ensure no sparse data is missed
            start_r, start_c = 1, 1
            end_r, end_c = sheet.max_row, sheet.max_column

            df, row_map, col_map = extract_table_with_coords(
                tmp_path, sheet_name, 
                start_row=start_r, start_col=start_c, 
                end_row=end_r, end_col=end_c
            )
            st.success("Table extracted and coordinates mapped successfully!")
            
            # Spreadsheet Preview Requirement
            st.subheader("Spreadsheet Preview")
            st.dataframe(df.head(50), use_container_width=True)
            
        except Exception as e:
            st.error(f"Error during extraction: {e}")
            st.stop()

    # --- Query Box ---
    st.markdown("---")
    st.subheader("Chat / Query")
    query = st.chat_input("Enter your query based on the spreadsheet...")

    if query:
        st.chat_message("user").write(query)
        
        if not google_api_key:
            st.error("Please provide a Gemini API Key in the sidebar.")
        else:
            with st.spinner("Generating parsing logic and querying via Gemini..."):
                code = generate_llm_code(df, query, google_api_key)
                
                if code:
                    try:
                        val, ref = execute_and_augment(code, df, row_map, col_map)
                        
                        if val is not None and ref is not None:
                            st.chat_message("assistant").markdown(f"**Value:** `{str(val)[:200]}`\n\n**Source Reference:** `{ref}`")
                            
                            # Log to JSON History Requirement
                            save_to_history(query, val, ref)
                            
                        else:
                            st.chat_message("assistant").warning("No matching row target found from execution.")
                    except Exception as e:
                        st.error(f"Execution Error running LLM Output: {e}")
