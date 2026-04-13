import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def auto_detect_table_boundaries(sheet):
    """
    Auto-detects the largest grid of relatively dense cells in the given sheet.
    Returns (start_row, start_col, end_row, end_col).
    """
    max_area = 0
    best_bounds = (None, None, None, None)
    
    for r in range(1, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            if sheet.cell(row=r, column=c).value is not None:
                # Find extent of contiguous data in this row (potential header)
                end_c = c
                while end_c <= sheet.max_column and sheet.cell(row=r, column=end_c).value is not None:
                    end_c += 1
                end_c -= 1
                
                end_r = r
                while end_r <= sheet.max_row:
                    row_has_data = False
                    for temp_c in range(c, end_c + 1):
                        if sheet.cell(row=end_r, column=temp_c).value is not None:
                            row_has_data = True
                            break
                    if not row_has_data:
                        break
                    end_r += 1
                end_r -= 1
                
                area = (end_r - r + 1) * (end_c - c + 1)
                
                if area > max_area:
                    max_area = area
                    best_bounds = (r, c, end_r, end_c)
                    
    if max_area == 0:
        raise ValueError("No table found in sheet.")
        
    return best_bounds


def extract_table_with_coords(file_path, sheet_name, start_row=None, start_col=None, end_row=None, end_col=None):
    """
    Extracts a rectangular grid from an Excel sheet.
    If boundaries are omitted, it auto-detects the largest contiguous table.
    Returns a Pandas DataFrame and a metadata dictionary tracking coordinates.
    """
    wb = load_workbook(file_path, data_only=True)
    sheet = wb[sheet_name]
    
    if None in (start_row, start_col, end_row, end_col):
        auto_start_r, auto_start_c, auto_end_r, auto_end_c = auto_detect_table_boundaries(sheet)
        start_row = start_row or auto_start_r
        start_col = start_col or auto_start_c
        end_row = end_row or auto_end_r
        end_col = end_col or auto_end_c
        print(f"Auto-detected table boundaries: Start({start_row}, {start_col}) to End({end_row}, {end_col})")
    
    data = []
    # This dictionary maps DataFrame row index -> Excel Sheet & Row
    row_mapping = {}
    
    # Treat the first row in the range as the header
    raw_headers = [str(sheet.cell(row=start_row, column=c).value) for c in range(start_col, end_col + 1)]
    headers = []
    seen = {}
    for h in raw_headers:
        if h not in seen:
            seen[h] = 1
            headers.append(h)
        else:
            headers.append(f"{h}_{seen[h]}")
            seen[h] += 1
    df_row_idx = 0
    for r in range(start_row + 1, end_row + 1):
        row_data = []
        for c in range(start_col, end_col + 1):
            row_data.append(sheet.cell(row=r, column=c).value)
        data.append(row_data)
        
        # Track where this dataframe row came from!
        row_mapping[df_row_idx] = {"sheet": sheet_name, "excel_row": r}
        df_row_idx += 1
        
    df = pd.DataFrame(data, columns=headers)
    
    # Map column headers to their original Excel Column Indices/Letters
    col_mapping = {
        col_name: {
            "excel_col_idx": start_col + i, 
            "excel_col_letter": get_column_letter(start_col + i)
        } 
        for i, col_name in enumerate(headers)
    }
    
    return df, row_mapping, col_mapping

if __name__ == "__main__":
    # Test script usage placeholder
    pass
